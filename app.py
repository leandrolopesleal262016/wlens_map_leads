import io
import os
import threading
import time
import webbrowser
from typing import Any
from urllib.parse import urlparse

import requests
from flask import Flask, abort, jsonify, render_template, request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REQUEST_TIMEOUT_SECONDS = int(os.getenv("REQUEST_TIMEOUT_SECONDS", "30"))
API_KEY_SENTINELS = {"", "YOUR_API_KEY_HERE"}
DETAILS_FIELDS = ",".join(
    [
        "address_components",
        "adr_address",
        "business_status",
        "formatted_address",
        "formatted_phone_number",
        "geometry",
        "international_phone_number",
        "name",
        "opening_hours",
        "place_id",
        "price_level",
        "rating",
        "type",
        "url",
        "user_ratings_total",
        "website",
    ]
)
HUBSPOT_EXPORT_COLUMNS = [
    ("Company name", lambda item: item.get("nome", "")),
    ("Company domain name", lambda item: item.get("dominio_empresa", "")),
    ("First name", lambda item: ""),
    ("Last name", lambda item: ""),
    ("Email", lambda item: ""),
    ("Phone number", lambda item: item.get("telefone_formatado", "")),
    ("Mobile phone number", lambda item: item.get("whatsapp", "")),
    ("Website URL", lambda item: item.get("site", "")),
    ("Address", lambda item: item.get("endereco", "")),
    ("City", lambda item: item.get("cidade", "")),
    ("State/Region", lambda item: item.get("estado", "")),
    ("Zip", lambda item: item.get("cep", "")),
    ("Country/Region", lambda item: item.get("pais", "")),
    ("Description", lambda item: item.get("descricao_lead", "")),
    ("Lead status", lambda item: ""),
    ("Lifecycle stage", lambda item: ""),
    ("Original source detail 1", lambda item: item.get("origem_importacao", "")),
    ("Google Search Query", lambda item: item.get("origem_busca", "")),
    ("Google Place ID", lambda item: item.get("place_id", "")),
    ("Google Maps URL", lambda item: item.get("google_maps_url", "")),
    ("Google Business Status", lambda item: item.get("status_negocio", "")),
    ("Google Types", lambda item: item.get("tipos_google", "")),
    ("Google Rating", lambda item: item.get("rating", "")),
    ("Google Reviews Count", lambda item: item.get("total_avaliacoes", "")),
    ("Google Open Now", lambda item: item.get("aberto_agora", "")),
    ("Google Price Level", lambda item: item.get("preco", "")),
    ("Neighborhood", lambda item: item.get("bairro", "")),
    ("Latitude", lambda item: item.get("lat", "")),
    ("Longitude", lambda item: item.get("lng", "")),
    ("Google International Phone", lambda item: item.get("telefone_internacional", "")),
]


def _load_api_key() -> str:
    env_key = os.getenv("GOOGLE_API_KEY", "").strip()
    if env_key:
        return env_key

    try:
        from chave import API_KEY as fallback_key
    except ImportError:
        return ""

    return (fallback_key or "").strip()


def _api_key_configured() -> bool:
    return API_KEY not in API_KEY_SENTINELS


def _get_google_error(payload: dict[str, Any]) -> str:
    status = payload.get("status", "")
    error_message = payload.get("error_message", "")

    if status in {"", "OK", "ZERO_RESULTS"}:
        return ""
    if error_message:
        return f"Google Maps API retornou {status}: {error_message}"
    return f"Google Maps API retornou {status}."


def _first_non_empty(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _normalize_phone(phone: str) -> str:
    return "".join(char for char in phone if char.isdigit())


def _domain_from_url(url: str) -> str:
    if not url:
        return ""

    parsed = urlparse(url if "://" in url else f"https://{url}")
    domain = (parsed.netloc or parsed.path).strip().lower()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain.split("/")[0]


def _address_component(
    components: list[dict[str, Any]], expected_types: tuple[str, ...]
) -> str:
    for component in components:
        types = component.get("types", [])
        if any(expected_type in types for expected_type in expected_types):
            return component.get("long_name", "")
    return ""


def _format_open_now(open_now: Any) -> str:
    if open_now is True:
        return "Sim"
    if open_now is False:
        return "Nao"
    return ""


def _lead_description() -> str:
    return (
        "Lead captado via Google Maps. "
        "Email nao e fornecido pela Google Places API."
    )


def _build_lead(
    termo: str, lugar: dict[str, Any], detalhes_resultado: dict[str, Any]
) -> dict[str, Any]:
    geometry = detalhes_resultado.get("geometry", {}).get("location", {})
    fallback_geometry = lugar.get("geometry", {}).get("location", {})
    lat = _first_non_empty(geometry.get("lat"), fallback_geometry.get("lat"))
    lng = _first_non_empty(geometry.get("lng"), fallback_geometry.get("lng"))

    components = detalhes_resultado.get("address_components", [])
    website = detalhes_resultado.get("website", "")
    formatted_phone = detalhes_resultado.get("formatted_phone_number", "")
    international_phone = detalhes_resultado.get("international_phone_number", "")
    primary_phone = _first_non_empty(
        international_phone, formatted_phone, "Nao informado"
    )
    phone_digits = _normalize_phone(
        _first_non_empty(international_phone, formatted_phone, "")
    )
    types = detalhes_resultado.get("types") or lugar.get("types") or []

    return {
        "nome": detalhes_resultado.get("name") or lugar.get("name", "Sem nome"),
        "numero": primary_phone,
        "telefone_formatado": _first_non_empty(formatted_phone, international_phone, ""),
        "telefone_internacional": international_phone,
        "whatsapp": phone_digits,
        "email": "",
        "site": website,
        "dominio_empresa": _domain_from_url(website),
        "endereco": detalhes_resultado.get("formatted_address")
        or lugar.get("formatted_address", "")
        or lugar.get("vicinity", ""),
        "bairro": _address_component(
            components, ("sublocality_level_1", "sublocality", "neighborhood")
        ),
        "cidade": _address_component(
            components, ("locality", "administrative_area_level_2")
        ),
        "estado": _address_component(components, ("administrative_area_level_1",)),
        "pais": _address_component(components, ("country",)),
        "cep": _address_component(components, ("postal_code",)),
        "google_maps_url": detalhes_resultado.get("url", ""),
        "status_negocio": detalhes_resultado.get("business_status")
        or lugar.get("business_status", ""),
        "tipos_google": ", ".join(types),
        "rating": detalhes_resultado.get("rating", ""),
        "total_avaliacoes": detalhes_resultado.get("user_ratings_total", ""),
        "aberto_agora": _format_open_now(
            detalhes_resultado.get("opening_hours", {}).get("open_now")
        ),
        "preco": detalhes_resultado.get("price_level", ""),
        "place_id": detalhes_resultado.get("place_id") or lugar.get("place_id", ""),
        "lat": lat,
        "lng": lng,
        "origem_busca": termo,
        "origem_importacao": "Google Maps",
        "descricao_lead": _lead_description(),
    }


def _buscar_leads(termo: str) -> list[dict[str, Any]]:
    resposta = requests.get(
        "https://maps.googleapis.com/maps/api/place/textsearch/json",
        params={"query": termo, "key": API_KEY},
        timeout=REQUEST_TIMEOUT_SECONDS,
    )
    resposta.raise_for_status()
    dados = resposta.json()

    erro = _get_google_error(dados)
    if erro:
        raise RuntimeError(erro)

    resultados = []
    for lugar in dados.get("results", []):
        detalhes_resultado: dict[str, Any] = {}
        place_id = lugar.get("place_id")
        if place_id:
            try:
                detalhes_resposta = requests.get(
                    "https://maps.googleapis.com/maps/api/place/details/json",
                    params={
                        "place_id": place_id,
                        "fields": DETAILS_FIELDS,
                        "key": API_KEY,
                    },
                    timeout=REQUEST_TIMEOUT_SECONDS,
                )
                detalhes_resposta.raise_for_status()
                detalhes = detalhes_resposta.json()

                detalhes_erro = _get_google_error(detalhes)
                if not detalhes_erro:
                    detalhes_resultado = detalhes.get("result", {})
            except (requests.RequestException, ValueError):
                pass

        lead = _build_lead(termo, lugar, detalhes_resultado)
        if lead.get("lat") in (None, "") or lead.get("lng") in (None, ""):
            continue
        resultados.append(lead)

    return resultados


def _criar_xlsx(linhas: list[dict[str, Any]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads HubSpot"

    headers = [header for header, _ in HUBSPOT_EXPORT_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in linhas or []:
        ws.append([resolver(item) for _, resolver in HUBSPOT_EXPORT_COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for index, column in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _browser_url(host: str, port: int) -> str:
    browser_host = "127.0.0.1" if host == "0.0.0.0" else host
    return f"http://{browser_host}:{port}"


app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"))
app.secret_key = os.getenv("SECRET_KEY", "change-me-in-production")
API_KEY = _load_api_key()


@app.route("/", methods=["GET", "POST"])
def index():
    # The results payload can easily exceed browser cookie limits.
    # Keep the Flask session empty so reverse proxies do not fail with 502.
    session.clear()
    resultados = []
    error_message = None
    termo = request.form.get("termo", "").strip()

    if request.method == "POST":
        if not termo:
            error_message = "Informe o termo da busca."
        elif not _api_key_configured():
            error_message = "Configure GOOGLE_API_KEY antes de usar a busca."
        else:
            try:
                resultados = _buscar_leads(termo)
            except requests.RequestException:
                error_message = "Nao foi possivel consultar o Google Maps agora."
            except ValueError:
                error_message = "Google Maps retornou uma resposta invalida."
            except RuntimeError as exc:
                error_message = str(exc)

    return render_template(
        "index.html",
        resultados=resultados,
        key=API_KEY if _api_key_configured() else "",
        maps_enabled=_api_key_configured(),
        error_message=error_message,
        termo=termo,
        download_label="Baixar Excel para HubSpot",
    )


@app.route("/baixar_xml", methods=["GET", "POST"])
def baixar_xml():
    """Mantem a rota legada, mas retorna um arquivo XLSX."""
    session.clear()
    if request.method == "POST":
        linhas = request.get_json(silent=True)
        if not isinstance(linhas, list):
            linhas = []
    else:
        linhas = []

    xlsx_io = _criar_xlsx(linhas)
    resp = send_file(
        xlsx_io,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="leads_hubspot.xlsx",
    )
    resp.headers["Content-Transfer-Encoding"] = "binary"
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "api_key_configured": _api_key_configured()})


@app.route("/sair", methods=["POST"])
def sair():
    if os.getenv("ALLOW_SHUTDOWN") != "1":
        abort(403)
    os._exit(0)


def abrir_navegador(host: str, port: int):
    time.sleep(1)
    webbrowser.open(_browser_url(host, port))


if __name__ == "__main__":
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG") == "1"

    if os.getenv("WLENS_OPEN_BROWSER") == "1":
        threading.Thread(target=abrir_navegador, args=(host, port), daemon=True).start()

    app.run(host=host, port=port, debug=debug)
